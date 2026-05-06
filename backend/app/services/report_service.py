"""Run report aggregation + Excel export.

Two consumers:
- ``GET /agent-runs/{id}/report`` (JSON) — frontend renders an in-app table
- ``GET /agent-runs/{id}/report.xlsx`` — streams an openpyxl workbook

Both share :func:`build_run_report`. Module / submodule grouping uses each
step row's frozen ``path_snapshot`` (``"Module > Submodule > Step"``)
rather than re-walking the live tc_node tree — that way the report
survives source-node deletion and edits, and avoids a join.
"""

from __future__ import annotations

import io
import logging
from collections import OrderedDict

from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.agent_run import AgentRun
from app.models.execution_step import ExecutionStep
from app.models.test_plan import TestPlan
from app.schemas.report import (
    ReportModuleRead,
    ReportPlanSummary,
    ReportRead,
    ReportRunSummary,
    ReportStepRead,
    ReportSubmoduleRead,
)

logger = logging.getLogger(__name__)


_PATH_SEP = " > "
_NONE_LABEL = "(none)"
# Truncate error excerpts that go into the issues column. Keeps the
# Excel cells readable + bounds the JSON payload size for big runs.
_ISSUE_EXCERPT_CHARS = 200
_ISSUES_PER_SUBMODULE = 5


# ── Per-step → report-row conversion ─────────────────────────────


def _step_to_report_row(step: ExecutionStep) -> ReportStepRead:
    """Flatten an ExecutionStep into a ReportStepRead, lifting AI flags
    out of ``details_json`` so the frontend doesn't have to re-parse JSON."""
    details = step.details_json or {}
    ai = details.get("ai_correction") if isinstance(details, dict) else None

    ai_helped = bool(ai) and step.status == "passed"
    ai_used_vision = (
        bool(ai)
        and isinstance(ai, dict)
        and bool(ai.get("used_vision"))
    )

    return ReportStepRead(
        id=step.id,
        tc_node_id=step.tc_node_id,
        ordinal=step.ordinal,
        title=step.title_snapshot,
        action_type=step.action_type_snapshot,
        target_hint=step.target_hint_snapshot,
        status=step.status,  # type: ignore[arg-type]
        duration_ms=step.duration_ms,
        screenshot_path=step.screenshot_path,
        error_message=step.error_message,
        narration=step.narration,
        ai_helped=ai_helped,
        ai_used_vision=ai_used_vision,
    )


def _split_path(path: str) -> tuple[str, str]:
    """Return (module_title, submodule_title) from a step's path_snapshot.

    Path is "Module > Submodule > Step". Falls back to ``(none)`` when
    the path is shallower than expected (e.g. a step generated without
    a parent submodule, or a path with extra separators).
    """
    parts = [p.strip() for p in (path or "").split(_PATH_SEP)]
    parts = [p for p in parts if p]
    if len(parts) >= 3:
        return parts[0], parts[1]
    if len(parts) == 2:
        return parts[0], _NONE_LABEL
    if len(parts) == 1:
        return parts[0], _NONE_LABEL
    return _NONE_LABEL, _NONE_LABEL


# ── Aggregation primitives ───────────────────────────────────────


def _zero_counts() -> dict[str, int]:
    return {"total": 0, "passed": 0, "failed": 0, "blocked": 0, "skipped": 0}


def _accumulate(counts: dict[str, int], status: str) -> None:
    counts["total"] += 1
    if status in ("passed", "failed", "blocked", "skipped"):
        counts[status] += 1


def _pcts(counts: dict[str, int]) -> tuple[float, float]:
    """Return (pass_pct, fail_pct) as floats in 0-100. Empty → 0/0.

    'fail_pct' counts only ``failed`` (not blocked/skipped) so the percent
    matches what users intuitively call "the failure rate".
    """
    total = counts["total"]
    if total == 0:
        return 0.0, 0.0
    passed = counts["passed"]
    failed = counts["failed"]
    return (
        round(100.0 * passed / total, 1),
        round(100.0 * failed / total, 1),
    )


# ── Top-level builder ────────────────────────────────────────────


def build_run_report(
    db: Session,
    run: AgentRun,
    *,
    excel_url: str,
) -> ReportRead:
    """Aggregate an execute run into a ReportRead tree.

    Modules and submodules preserve insertion order — the order steps
    were *executed* (DFS of the TC tree), so the report rows read
    top-to-bottom matching what the user saw on the timeline.
    """
    # Pull plan info if still around (SET NULL on plan delete)
    plan_summary: ReportPlanSummary | None = None
    if run.plan_id is not None:
        plan = db.get(TestPlan, run.plan_id)
        if plan is not None:
            plan_summary = ReportPlanSummary(
                id=plan.id,
                name=plan.name,
                target_url=plan.target_url or "",
                scope=list(plan.scope or []),
            )

    rows = list(
        db.scalars(
            select(ExecutionStep)
            .where(ExecutionStep.run_id == run.id)
            .order_by(ExecutionStep.ordinal),
        ),
    )

    # Two-level OrderedDict keyed by titles for stable execution-order
    # output. Inner value: {counts: dict, issues: list[str], steps: list}.
    by_module: OrderedDict[str, OrderedDict[str, dict]] = OrderedDict()
    run_counts = _zero_counts()

    for step in rows:
        module_title, submodule_title = _split_path(step.path_snapshot or "")
        mod_dict = by_module.setdefault(module_title, OrderedDict())
        sub_dict = mod_dict.setdefault(
            submodule_title,
            {"counts": _zero_counts(), "issues": [], "steps": []},
        )

        _accumulate(sub_dict["counts"], step.status)
        _accumulate(run_counts, step.status)

        if step.status == "failed" and step.error_message:
            excerpt = step.error_message.strip()[:_ISSUE_EXCERPT_CHARS]
            if excerpt and excerpt not in sub_dict["issues"]:
                if len(sub_dict["issues"]) < _ISSUES_PER_SUBMODULE:
                    sub_dict["issues"].append(excerpt)

        sub_dict["steps"].append(_step_to_report_row(step))

    # Build the report tree from the dict-of-dicts
    modules: list[ReportModuleRead] = []
    for mod_title, sub_map in by_module.items():
        mod_counts = _zero_counts()
        submodules: list[ReportSubmoduleRead] = []
        for sub_title, payload in sub_map.items():
            for k, v in payload["counts"].items():
                mod_counts[k] += v
            sub_pass, sub_fail = _pcts(payload["counts"])
            submodules.append(
                ReportSubmoduleRead(
                    title=sub_title,
                    total=payload["counts"]["total"],
                    passed=payload["counts"]["passed"],
                    failed=payload["counts"]["failed"],
                    blocked=payload["counts"]["blocked"],
                    skipped=payload["counts"]["skipped"],
                    pass_pct=sub_pass,
                    fail_pct=sub_fail,
                    issues=list(payload["issues"]),
                    steps=list(payload["steps"]),
                ),
            )
        mod_pass, mod_fail = _pcts(mod_counts)
        modules.append(
            ReportModuleRead(
                title=mod_title,
                total=mod_counts["total"],
                passed=mod_counts["passed"],
                failed=mod_counts["failed"],
                blocked=mod_counts["blocked"],
                skipped=mod_counts["skipped"],
                pass_pct=mod_pass,
                fail_pct=mod_fail,
                submodules=submodules,
            ),
        )

    output = run.output_summary_json or {}
    run_pass, run_fail = _pcts(run_counts)
    summary = ReportRunSummary(
        id=run.id,
        status=run.status,  # type: ignore[arg-type]
        started_at=run.started_at,
        completed_at=run.completed_at,
        duration_ms=output.get("duration_ms")
        if isinstance(output.get("duration_ms"), int) else None,
        total_steps=run_counts["total"],
        passed=run_counts["passed"],
        failed=run_counts["failed"],
        blocked=run_counts["blocked"],
        skipped=run_counts["skipped"],
        pass_pct=run_pass,
        fail_pct=run_fail,
        llm_input_tokens=output.get("llm_input_tokens")
        if isinstance(output.get("llm_input_tokens"), int) else None,
        llm_output_tokens=output.get("llm_output_tokens")
        if isinstance(output.get("llm_output_tokens"), int) else None,
        ai_calls=int(output.get("ai_calls") or 0),
        ai_vision_calls=int(output.get("ai_vision_calls") or 0),
    )

    return ReportRead(
        run=summary,
        plan=plan_summary,
        modules=modules,
        excel_download_url=excel_url,
    )


# ── Excel export ─────────────────────────────────────────────────


def build_excel_workbook(report: ReportRead) -> bytes:
    """Generate an xlsx workbook with two sheets:

    - **Summary**: run-level numbers + plan info + AI cost
    - **Results**: one row per step with module/submodule/title/status/etc.

    Returns bytes ready to stream as an HTTP response.
    """
    # Local import — openpyxl is a heavy-ish dep (~5MB); only load it on
    # the export path so the rest of the API stays lean.
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # ── Summary sheet ────────────────────────────────────────────
    s = wb.active
    s.title = "Summary"
    bold = Font(bold=True)
    accent = Font(bold=True, color="FFFFFF")
    accent_fill = PatternFill("solid", fgColor="2A2F3A")

    rows: list[tuple[str, object]] = [
        ("Run", f"#{report.run.id}"),
        ("Status", report.run.status),
        ("Started", report.run.started_at.isoformat()
            if report.run.started_at else ""),
        ("Completed", report.run.completed_at.isoformat()
            if report.run.completed_at else ""),
        ("Duration (ms)", report.run.duration_ms or ""),
    ]
    if report.plan:
        rows.append(("Plan", report.plan.name))
        rows.append(("Target URL", report.plan.target_url))
        if report.plan.scope:
            rows.append(("Scope", ", ".join(report.plan.scope)))

    rows.extend([
        ("", ""),  # spacer
        ("Total steps", report.run.total_steps),
        ("Passed", report.run.passed),
        ("Failed", report.run.failed),
        ("Blocked", report.run.blocked),
        ("Skipped", report.run.skipped),
        ("Pass %", report.run.pass_pct),
        ("Fail %", report.run.fail_pct),
    ])
    if report.run.ai_calls > 0:
        rows.extend([
            ("", ""),
            ("AI assist calls", report.run.ai_calls),
            ("AI vision calls", report.run.ai_vision_calls),
            ("LLM input tokens",  report.run.llm_input_tokens or 0),
            ("LLM output tokens", report.run.llm_output_tokens or 0),
        ])

    for r_idx, (label, value) in enumerate(rows, start=1):
        s.cell(row=r_idx, column=1, value=label).font = bold
        s.cell(row=r_idx, column=2, value=value)
    s.column_dimensions["A"].width = 22
    s.column_dimensions["B"].width = 60

    # ── Results sheet ────────────────────────────────────────────
    r = wb.create_sheet(title="Results")
    headers = [
        "Module",
        "Submodule",
        "#",
        "Title",
        "Action",
        "Status",
        "Duration (ms)",
        "AI helped",
        "Used vision",
        "Issues",
    ]
    for col_idx, h in enumerate(headers, start=1):
        cell = r.cell(row=1, column=col_idx, value=h)
        cell.font = accent
        cell.fill = accent_fill
        cell.alignment = Alignment(horizontal="left", vertical="center")

    status_fills = {
        "passed":  PatternFill("solid", fgColor="DCFCE7"),  # light green
        "failed":  PatternFill("solid", fgColor="FEE2E2"),  # light red
        "blocked": PatternFill("solid", fgColor="FEF9C3"),  # light yellow
        "skipped": PatternFill("solid", fgColor="F1F5F9"),  # light gray
    }

    row_idx = 2
    for module in report.modules:
        for sub in module.submodules:
            for step in sub.steps:
                values = [
                    module.title,
                    sub.title,
                    step.ordinal + 1,
                    step.title,
                    step.action_type or "",
                    step.status,
                    step.duration_ms or 0,
                    "yes" if step.ai_helped else "",
                    "yes" if step.ai_used_vision else "",
                    (step.error_message or "")[:300],
                ]
                for col_idx, v in enumerate(values, start=1):
                    cell = r.cell(row=row_idx, column=col_idx, value=v)
                    if col_idx == 6:  # Status column
                        fill = status_fills.get(step.status)
                        if fill:
                            cell.fill = fill
                row_idx += 1

    # Column widths — generous for titles, narrow for numerics
    widths = [22, 28, 5, 38, 10, 9, 12, 9, 10, 60]
    for col_idx, w in enumerate(widths, start=1):
        r.column_dimensions[get_column_letter(col_idx)].width = w

    # Freeze header row
    r.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
